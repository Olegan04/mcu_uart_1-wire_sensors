import sys
import serial
import serial.tools.list_ports
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import threading
import time
import re
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import subprocess
import platform

class DS18B20Monitor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.serial_port = None
        self.is_connected = False
        self.reading_thread = None
        self.stop_thread = False
        
        # Флаг ошибки чтения
        self.read_error_occurred = False
        # Флаг режима переподключения
        self.reconnect_mode = False
        
        # Таймер для мигания индикатора
        self.indicator_timer = QTimer()
        self.indicator_timer.timeout.connect(self.update_indicator)
        self.indicator_state = False  # Текущее состояние индикатора (вкл/выкл)
        
        # Данные датчиков
        self.sensor_data = {
            0: {"temp": "---", "res": "12", "working": True, "last_saved_temp": None},
            1: {"temp": "---", "res": "12", "working": True, "last_saved_temp": None}
        }
        
        # Для записи в Excel
        self.log_data = []
        self.excel_file = "temperature_log.xlsx"
        
        self.init_ui()
        self.scan_ports()
        
        # Создаем/открываем Excel файл при запуске
        self.open_or_create_excel()
        
    def init_ui(self):
        # Настройка главного окна
        self.setWindowTitle("DS18B20 Monitor - STM32")
        
        # Открываем во весь экран
        self.showFullScreen()
        
        # Центральный виджет
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(5)
        
        # 1. Заголовок и кнопка закрытия
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        title = QLabel("Мониторинг температуры DS18B20")
        title.setStyleSheet("""
            font-size: 38px;
            font-weight: bold;
            color: #2c3e50;
        """)
        
        # Кнопка закрытия
        self.close_btn = QPushButton("✕ Закрыть")
        self.close_btn.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                font-weight: bold;
                padding: 10px 20px;
                background-color: #e74c3c;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.close_btn.clicked.connect(self.close)
        
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(self.close_btn)
        
        layout.addWidget(header_widget)
        
        # 2. Панель подключения
        conn_frame = QFrame()
        conn_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        conn_frame.setLineWidth(2)
        conn_layout = QHBoxLayout(conn_frame)
        conn_layout.setSpacing(5)
        
        # Выбор порта
        port_label = QLabel("COM-порт:")
        port_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        self.port_combo = QComboBox()
        self.port_combo.setMinimumWidth(200)
        self.port_combo.setStyleSheet("font-size: 20px; padding: 8px;")
        
        # # Выбор скорости
        # baud_label = QLabel("Скорость:")
        # baud_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        # self.baud_combo = QLabel("9600")
        # self.baud_combo.setStyleSheet("font-size: 30x; padding: 8px;")
        
        # Кнопки
        self.refresh_btn = QPushButton("🔄 Обновить")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                font-weight: bold;
                padding: 10px 15px;
                background-color: #3498db;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.refresh_btn.clicked.connect(self.scan_ports)
        
        # Кнопка подключения/отключения/переподключения
        self.connect_btn = QPushButton("🔗 Подключиться")
        self.connect_btn.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                font-weight: bold;
                padding: 10px 15px;
                background-color: #2ecc71;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        self.connect_btn.clicked.connect(self.toggle_connection)

        # Индикатор подключения (добавлен справа)
        self.indicator_label = QLabel("●")
        self.indicator_label.setAlignment(Qt.AlignCenter)
        self.indicator_label.setStyleSheet("""
            QLabel {
                font-size: 360px;
                font-weight: bold;
                color: #e74c3c;
                padding: 0px;
                border-radius: 20px;
                background-color: #ecf0f1;
            }
        """)
        
        # Добавляем элементы
        conn_layout.addWidget(port_label)
        conn_layout.addWidget(self.port_combo)
        conn_layout.addStretch()  # Добавляем растягивающий элемент
        conn_layout.addWidget(self.refresh_btn)
        conn_layout.addWidget(self.connect_btn)
        conn_layout.addWidget(self.indicator_label)
        
        layout.addWidget(conn_frame)
        
        # 3. Отображение температуры
        temp_frame = QFrame()
        temp_layout = QHBoxLayout(temp_frame)
        temp_layout.setSpacing(20)
        
        # Датчик 1
        self.sensor1_frame = QFrame()
        self.sensor1_frame.setFrameStyle(QFrame.Box | QFrame.Raised)
        self.sensor1_frame.setLineWidth(3)
        self.sensor1_frame.setMinimumHeight(250)
        sensor1_layout = QVBoxLayout(self.sensor1_frame)
        
        sensor1_title = QLabel("🌡️ ДАТЧИК 1")
        sensor1_title.setAlignment(Qt.AlignCenter)
        sensor1_title.setStyleSheet("""
            font-size: 34px;
            font-weight: bold;
            color: #3498db;
            padding: 10px;
        """)
        
        self.sensor1_temp = QLabel("--- °C")
        self.sensor1_temp.setAlignment(Qt.AlignCenter)
        self.sensor1_temp.setStyleSheet("""
            font-size: 90px;
            font-weight: bold;
            color: #3498db;
            padding: 20px 0;
        """)
        
        self.sensor1_status = QLabel("Статус: ожидание...")
        self.sensor1_status.setAlignment(Qt.AlignCenter)
        self.sensor1_status.setStyleSheet("""
            font-size: 34px;
            color: #000000;
            padding: 10px;
        """)
        
        sensor1_layout.addWidget(sensor1_title)
        sensor1_layout.addWidget(self.sensor1_temp)
        sensor1_layout.addWidget(self.sensor1_status)
        
        # Датчик 2
        self.sensor2_frame = QFrame()
        self.sensor2_frame.setFrameStyle(QFrame.Box | QFrame.Raised)
        self.sensor2_frame.setLineWidth(3)
        self.sensor2_frame.setMinimumHeight(250)
        sensor2_layout = QVBoxLayout(self.sensor2_frame)
        
        sensor2_title = QLabel("🌡️ ДАТЧИК 2")
        sensor2_title.setAlignment(Qt.AlignCenter)
        sensor2_title.setStyleSheet("""
            font-size: 34px;
            font-weight: bold;
            color: #e74c3c;
            padding: 10px;
        """)
        
        self.sensor2_temp = QLabel("--- °C")
        self.sensor2_temp.setAlignment(Qt.AlignCenter)
        self.sensor2_temp.setStyleSheet("""
            font-size: 90px;
            font-weight: bold;
            color: #e74c3c;
            padding: 20px 0;
        """)
        
        self.sensor2_status = QLabel("Статус: ожидание...")
        self.sensor2_status.setAlignment(Qt.AlignCenter)
        self.sensor2_status.setStyleSheet("""
            font-size: 34px;
            color: #000000;
            padding: 10px;
        """)
        
        sensor2_layout.addWidget(sensor2_title)
        sensor2_layout.addWidget(self.sensor2_temp)
        sensor2_layout.addWidget(self.sensor2_status)
        
        temp_layout.addWidget(self.sensor1_frame)
        temp_layout.addWidget(self.sensor2_frame)
        layout.addWidget(temp_frame)
        
        # 4. Радиокнопки для выбора разрешения
        resolution_frame = QFrame()
        resolution_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        resolution_layout = QHBoxLayout(resolution_frame)
        resolution_layout.setSpacing(20)
        
        # Группа для датчика 1
        sensor1_res_group = QGroupBox("Разрешение Датчика 1")
        sensor1_res_group.setStyleSheet("font-size: 30px")
        sensor1_res_layout = QHBoxLayout()
        
        self.sensor1_res_buttons = {}
        resolutions = [("9 бит", "9"), ("10 бит", "10"), ("11 бит", "11"), ("12 бит", "12")]
        
        for text, value in resolutions:
            btn = QRadioButton(text)
            btn.setProperty("sensor", 0)
            btn.setProperty("value", value)
            if value == "12":
                btn.setChecked(True)
            btn.toggled.connect(self.on_resolution_changed)
            btn.setEnabled(False)
            self.sensor1_res_buttons[value] = btn
            sensor1_res_layout.addWidget(btn)
        
        sensor1_res_group.setLayout(sensor1_res_layout)
        
        # Группа для датчика 2
        sensor2_res_group = QGroupBox("Разрешение Датчика 2")
        sensor2_res_group.setStyleSheet("font-size: 30px")
        sensor2_res_layout = QHBoxLayout()
        
        self.sensor2_res_buttons = {}
        for text, value in resolutions:
            btn = QRadioButton(text)
            btn.setProperty("sensor", 1)
            btn.setProperty("value", value)
            if value == "12":
                btn.setChecked(True)
            btn.toggled.connect(self.on_resolution_changed)
            btn.setEnabled(False)
            self.sensor2_res_buttons[value] = btn
            sensor2_res_layout.addWidget(btn)
        
        sensor2_res_group.setLayout(sensor2_res_layout)
        
        resolution_layout.addWidget(sensor1_res_group)
        resolution_layout.addWidget(sensor2_res_group)
        layout.addWidget(resolution_frame)
        
        # 5. Информация о записи в Excel
        self.excel_frame = QFrame()
        self.excel_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        excel_layout = QHBoxLayout(self.excel_frame)
        
        self.excel_label = QLabel("📁 Файл Excel: temperature_log.xlsx")
        self.excel_label.setStyleSheet("font-size: 25px; color: #2c3e50;")
        
        self.open_excel_btn = QPushButton("📂 Открыть Excel")
        self.open_excel_btn.setStyleSheet("""
            QPushButton {
                font-size: 25px;
                padding: 8px 15px;
                background-color: #2ecc71;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        self.open_excel_btn.clicked.connect(self.open_excel_file)
        
        excel_layout.addWidget(self.excel_label)
        excel_layout.addStretch()
        excel_layout.addWidget(self.open_excel_btn)
        
        layout.addWidget(self.excel_frame)
        
        # 7. Статус бар
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Готов к работе")
        self.status_bar.setStyleSheet("""
            QStatusBar {
                background-color: #34495e;
                color: white;
                font-weight: bold;
                font-size: 20px;
            }
        """)
        
    def update_indicator(self):
        """Обновление состояния индикатора подключения"""
        if self.is_connected:
            # Мигание при подключении
            self.indicator_state = not self.indicator_state
            if self.indicator_state:
                # Включенное состояние (зеленый)
                self.indicator_label.setStyleSheet("""
                    QLabel {
                        font-size: 360px;
                        font-weight: bold;
                        color: #2ecc71;
                        border-radius: 10px;
                        background-color: #ecf0f1;
                    }
                """)
            else:
                # Выключенное состояние (серый)
                self.indicator_label.setStyleSheet("""
                    QLabel {
                        font-size: 360px;
                        font-weight: bold;
                        color: #95a5a6;
                        border-radius: 10px;
                        background-color: #ecf0f1;
                    }
                """)
        else:
            # Постоянный красный при отключении
            self.indicator_state = False
            self.indicator_label.setStyleSheet("""
                QLabel {
                    font-size: 360px;
                    font-weight: bold;
                    color: #e74c3c;
                    border-radius: 10px;
                    background-color: #ecf0f1;
                }
            """)
    
    def start_indicator_blink(self):
        """Запуск мигания индикатора"""
        if not self.indicator_timer.isActive():
            self.indicator_timer.start(500)  # Мигание каждые 500 мс
    
    def stop_indicator_blink(self):
        """Остановка мигания индикатора"""
        if self.indicator_timer.isActive():
            self.indicator_timer.stop()
        self.update_indicator()  # Обновляем до статичного состояния
    
    def open_or_create_excel(self):
        """Открытие существующего или создание нового Excel файла"""
        try:
            if os.path.exists(self.excel_file):
                # Читаем существующий файл
                existing_df = pd.read_excel(self.excel_file)
                self.status_bar.showMessage(f"Загружен существующий файл Excel. Всего записей: {len(existing_df)}", 3000)
                
                # Получаем последние значения температуры для каждого датчика
                if not existing_df.empty:
                    # Ищем последние валидные значения температуры для каждого датчика
                    for sensor_num in [0, 1]:
                        col_name = f'Датчик {sensor_num + 1} Температура (°C)'
                        if col_name in existing_df.columns:
                            # Ищем последнее не-ERROR значение
                            valid_values = existing_df[existing_df[col_name] != 'ERROR'][col_name]
                            if not valid_values.empty:
                                last_value = valid_values.iloc[-1]
                                try:
                                    # Пробуем преобразовать в float
                                    float_val = float(last_value)
                                    self.sensor_data[sensor_num]["last_saved_temp"] = float_val
                                except:
                                    pass
            else:
                # Создаем новый файл
                self.create_excel_file()
                
        except Exception as e:
            self.status_bar.showMessage(f"Ошибка работы с Excel файлом: {str(e)}", 5000)
            # Пробуем создать новый файл
            self.create_excel_file()
    
    def create_excel_file(self):
        """Создание нового Excel файла с заголовками"""
        try:
            # Создаем Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Температура"
            
            # Заголовки
            headers = ['Время', 'Датчик 1 Температура (°C)', 'Датчик 1 Статус', 
                      'Датчик 1 Разрешение (бит)', 'Датчик 2 Температура (°C)', 
                      'Датчик 2 Статус', 'Датчик 2 Разрешение (бит)']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[chr(64 + col)].width = 25
            
            wb.save(self.excel_file)
            self.status_bar.showMessage(f"Создан новый файл Excel: {self.excel_file}", 3000)
            
        except Exception as e:
            self.status_bar.showMessage(f"Ошибка создания Excel файла: {str(e)}", 5000)
    
    def open_excel_file(self):
        """Открытие Excel файла в системе"""
        try:
            if os.path.exists(self.excel_file):
                system = platform.system()
                if system == "Windows":
                    os.startfile(self.excel_file)
                elif system == "Darwin":  # macOS
                    subprocess.run(["open", self.excel_file])
                else:  # Linux
                    subprocess.run(["xdg-open", self.excel_file])
                self.status_bar.showMessage(f"Открыт файл Excel: {self.excel_file}", 3000)
            else:
                self.status_bar.showMessage("Файл Excel не найден! Создаем новый...", 3000)
                self.create_excel_file()
                
        except Exception as e:
            self.status_bar.showMessage(f"Ошибка открытия файла: {str(e)}", 5000)
    
    def save_to_excel_if_changed(self):
        """Сохраняет данные в Excel только если есть изменения"""
        try:
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Получаем данные датчиков
            temp1 = self.sensor_data[0]["temp"]
            temp2 = self.sensor_data[1]["temp"]
            working1 = self.sensor_data[0]["working"]
            working2 = self.sensor_data[1]["working"]
            
            # Определяем статус
            status1 = "ERROR" if not working1 else "OK"
            status2 = "ERROR" if not working2 else "OK"
            
            # Если датчик не работает, ставим "ERROR" вместо температуры
            if not working1:
                temp1 = "ERROR"
            if not working2:
                temp2 = "ERROR"
            
            # Проверяем, изменилась ли температура по сравнению с последней сохраненной
            temp1_changed = False
            temp2_changed = False
            
            try:
                if temp1 != "ERROR" and temp1 != "---":
                    temp1_float = float(temp1)
                    last_temp1 = self.sensor_data[0]["last_saved_temp"]
                    if last_temp1 is None or abs(temp1_float - last_temp1) > 0.01:  # Порог 0.01°C
                        temp1_changed = True
                        self.sensor_data[0]["last_saved_temp"] = temp1_float
            except:
                # Если не удалось преобразовать в float, считаем что данные изменились
                temp1_changed = True
            
            try:
                if temp2 != "ERROR" and temp2 != "---":
                    temp2_float = float(temp2)
                    last_temp2 = self.sensor_data[1]["last_saved_temp"]
                    if last_temp2 is None or abs(temp2_float - last_temp2) > 0.01:  # Порог 0.01°C
                        temp2_changed = True
                        self.sensor_data[1]["last_saved_temp"] = temp2_float
            except:
                temp2_changed = True
            
            # Также сохраняем если изменился статус работы датчика
            status_changed = False
            # Здесь можно добавить логику отслеживания изменения статуса
            
            # Сохраняем только если есть изменения
            if temp1_changed or temp2_changed or not working1 or not working2:
                new_row = {
                    'Время': current_time,
                    'Датчик 1 Температура (°C)': temp1,
                    'Датчик 1 Статус': status1,
                    'Датчик 1 Разрешение (бит)': self.sensor_data[0]["res"],
                    'Датчик 2 Температура (°C)': temp2,
                    'Датчик 2 Статус': status2,
                    'Датчик 2 Разрешение (бит)': self.sensor_data[1]["res"]
                }
                
                # Читаем существующие данные
                try:
                    existing_df = pd.read_excel(self.excel_file)
                except:
                    existing_df = pd.DataFrame()
                
                # Добавляем новую строку
                new_df = pd.DataFrame([new_row])
                
                if not existing_df.empty:
                    final_df = pd.concat([existing_df, new_df], ignore_index=True)
                else:
                    final_df = new_df
                
                # Сохраняем
                final_df.to_excel(self.excel_file, index=False)
            
                    
        except Exception as e:
            self.status_bar.showMessage(f"Ошибка сохранения в Excel: {str(e)}", 5000)
    
    def scan_ports(self):
        """Сканирование портов с автоопределением STM32"""
        self.port_combo.clear()
        ports = serial.tools.list_ports.comports()
        
        stm_port_found = None
        
        for port in ports:
            desc = port.description.lower()
            # Проверяем, это ли STM32
            is_stm = any(keyword in desc for keyword in ['stm', 'st-link', 'cmsis-dap', 'composite device'])
            
            display_text = f"{port.device} - {port.description}"
            if is_stm:
                display_text += " [STM32]"
                stm_port_found = port
            
            self.port_combo.addItem(display_text, port.device)
        
        # Автовыбор STM32 порта
        if stm_port_found:
            for i in range(self.port_combo.count()):
                if "[STM32]" in self.port_combo.itemText(i):
                    self.port_combo.setCurrentIndex(i)
                    self.status_bar.showMessage(f"Автовыбран порт STM32: {stm_port_found.device}", 3000)
                    break
        
        if ports:
            self.status_bar.showMessage(f"Найдено портов: {len(ports)}", 3000)
            self.connect_btn.setEnabled(True)
        else:
            self.status_bar.showMessage("Порты не найдены", 5000)
            self.connect_btn.setEnabled(False)
    
    def on_resolution_changed(self):
        """Обработка изменения """
        sender = self.sender()
        if sender.isChecked():
            sensor_num = sender.property("sensor")
            resolution = sender.property("value")
            
            # Обновляем данные
            self.sensor_data[sensor_num]["res"] = resolution
            
            # Отправляем команду на STM32
            self.send_resolution_command(sensor_num, resolution)
            
            self.status_bar.showMessage(f"Датчик {sensor_num + 1}: установлено разрешение {resolution} бит", 3000)
            self.update_display()
            
            # Сохраняем изменение разрешения
            self.save_to_excel_if_changed()
    
    def send_resolution_command(self, sensor_num, resolution):
        """Отправка команды для изменения разрешения"""
        # Специальные символы для команд:
        # Датчик 0: 
        #   9 бит = 'a', 10 бит = 'b', 11 бит = 'c', 12 бит = 'd'
        # Датчик 1:
        #   9 бит = 'e', 10 бит = 'f', 11 бит = 'g', 12 бит = 'h'
        
        command_map = {
            0: {"9": 'a', "10": 'b', "11": 'c', "12": 'd'},
            1: {"9": 'e', "10": 'f', "11": 'g', "12": 'h'}
        }
        
        if sensor_num in command_map and resolution in command_map[sensor_num]:
            cmd = command_map[sensor_num][resolution]
            self.send_command(cmd)
    
    def toggle_connection(self):
        """Подключение/отключение/переподключение"""
        if self.reconnect_mode:
            # Режим переподключения
            self.reconnect()
        elif not self.is_connected:
            # Обычное подключение
            self.connect()
        else:
            # Обычное отключение
            self.disconnect()
    
    def connect(self):
        """Подключение к порту"""
        if self.port_combo.currentIndex() < 0:
            self.status_bar.showMessage("Ошибка: не выбран порт!", 5000)
            return
        
        port = self.port_combo.currentData()
        baud = 9600
        
        try:
            self.serial_port = serial.Serial(port, baud, timeout=1)
            self.is_connected = True
            self.read_error_occurred = False
            self.reconnect_mode = False
            
            # Обновление интерфейса
            self.connect_btn.setText("🔌 Отключиться")
            self.connect_btn.setStyleSheet("""
                QPushButton {
                    font-size: 20px;
                    font-weight: bold;
                    padding: 10px 15px;
                    background-color: #e74c3c;
                    color: white;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            
            # Запускаем мигание индикатора
            self.start_indicator_blink()
            
            # Активация радиокнопок
            for btn in self.sensor1_res_buttons.values():
                btn.setEnabled(True)
            for btn in self.sensor2_res_buttons.values():
                btn.setEnabled(True)
            
            # Запуск потока чтения
            self.stop_thread = False
            self.reading_thread = threading.Thread(target=self.read_serial)
            self.reading_thread.daemon = True
            self.reading_thread.start()
            
            self.status_bar.showMessage(f"Успешно подключено к {port} ({baud} бод)")
            
        except Exception as e:
            self.status_bar.showMessage(f"Ошибка подключения: {str(e)}", 5000)
    
    def disconnect(self):
        """Отключение от порта"""
        self.stop_thread = True
        self.read_error_occurred = False
        self.reconnect_mode = False
        
        if self.reading_thread:
            self.reading_thread.join(timeout=0.5)
        
        if self.serial_port:
            self.serial_port.close()
        
        self.is_connected = False
        self.connect_btn.setText("🔗 Подключиться")
        self.connect_btn.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                font-weight: bold;
                padding: 10px 15px;
                background-color: #2ecc71;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        
        # Останавливаем мигание индикатора
        self.stop_indicator_blink()
        
        # Деактивация радиокнопок
        for btn in self.sensor1_res_buttons.values():
            btn.setEnabled(False)
        for btn in self.sensor2_res_buttons.values():
            btn.setEnabled(False)
        
        # Сброс отображения
        self.sensor1_temp.setText("--- °C")
        self.sensor2_temp.setText("--- °C")
        self.sensor1_status.setText("Статус: отключен")
        self.sensor2_status.setText("Статус: отключен")
        
        self.status_bar.showMessage("Отключено от порта")
        
        # Сохраняем данные при отключении
        self.save_to_excel_if_changed()
    
    def reconnect(self):
        """Переподключение к порту после потери связи"""
        # Сначала отключаемся
        self.stop_thread = True
        
        if self.reading_thread:
            self.reading_thread.join(timeout=0.5)
        
        if self.serial_port:
            try:
                self.serial_port.close()
            except:
                pass
        
        # Затем пытаемся подключиться заново
        self.status_bar.showMessage("Попытка переподключения...")
        
        # Сбрасываем флаги
        self.read_error_occurred = False
        self.reconnect_mode = False
        
        # Пытаемся подключиться
        if self.port_combo.currentIndex() < 0:
            self.status_bar.showMessage("Ошибка: не выбран порт!", 5000)
            self.connect_btn.setText("🔗 Подключиться")
            self.connect_btn.setStyleSheet("""
                QPushButton {
                    font-size: 20px;
                    font-weight: bold;
                    padding: 10px 15px;
                    background-color: #2ecc71;
                    color: white;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #27ae60;
                }
            """)
            return
        
        port = self.port_combo.currentData()
        baud = 9600
        
        try:
            self.serial_port = serial.Serial(port, baud, timeout=1)
            self.is_connected = True
            
            # Обновление интерфейса
            self.connect_btn.setText("🔌 Отключиться")
            self.connect_btn.setStyleSheet("""
                QPushButton {
                    font-size: 20px;
                    font-weight: bold;
                    padding: 10px 15px;
                    background-color: #e74c3c;
                    color: white;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            
            # Запускаем мигание индикатора
            self.start_indicator_blink()
            
            # Активация радиокнопок
            for btn in self.sensor1_res_buttons.values():
                btn.setEnabled(True)
            for btn in self.sensor2_res_buttons.values():
                btn.setEnabled(True)
            
            # Запуск потока чтения
            self.stop_thread = False
            self.reading_thread = threading.Thread(target=self.read_serial)
            self.reading_thread.daemon = True
            self.reading_thread.start()
            
            self.status_bar.showMessage(f"✅ Успешно переподключено к {port} ({baud} бод)")
            
        except Exception as e:
            self.status_bar.showMessage(f"❌ Ошибка переподключения: {str(e)}", 5000)
            self.is_connected = False
            self.connect_btn.setText("🔄 Переподключиться")
            self.connect_btn.setStyleSheet("""
                QPushButton {
                    font-size: 20px;
                    font-weight: bold;
                    padding: 10px 15px;
                    background-color: #f39c12;
                    color: white;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #e67e22;
                }
            """)
            self.reconnect_mode = True
    
    def read_serial(self):
        """Чтение данных из порта"""
        buffer = ""
        while not self.stop_thread and self.serial_port:
            try:
                if self.serial_port.in_waiting:
                    data = self.serial_port.read(self.serial_port.in_waiting).decode('utf-8', 'ignore')
                    buffer += data
                    
                    while '\n' in buffer:
                        line, buffer = buffer.split('\n', 1)
                        line = line.strip()
                        if line:
                            QMetaObject.invokeMethod(self, "process_line", 
                                                    Qt.QueuedConnection,
                                                    Q_ARG(str, line))
                
                time.sleep(0.01)
                
            except Exception as e:
                if not self.stop_thread:
                    # Устанавливаем флаг ошибки чтения
                    self.read_error_occurred = True
                    
                    # Включаем режим переподключения
                    self.reconnect_mode = True

                    self.is_connected = False
                    
                    # Обновляем кнопку для переподключения
                    QMetaObject.invokeMethod(self, "update_button_for_reconnect", 
                                            Qt.QueuedConnection)
                    
                    # Обновляем статус датчиков при ошибке чтения
                    QMetaObject.invokeMethod(self, "handle_read_error", 
                                            Qt.QueuedConnection)
                    
                    # Выводим ошибку в статус бар
                    error_msg = f"Ошибка чтения: потеря связи с устройством"
                    QMetaObject.invokeMethod(self.status_bar, "showMessage",
                                            Qt.QueuedConnection,
                                            Q_ARG(str, error_msg),
                                            Q_ARG(int, 5000))
                    break
    
    @pyqtSlot()
    def update_button_for_reconnect(self):
        """Обновление кнопки для режима переподключения"""
        self.connect_btn.setText("🔄 Переподключиться")
        self.connect_btn.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                font-weight: bold;
                padding: 10px 15px;
                background-color: #f39c12;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #e67e22;
            }
        """)
    
    @pyqtSlot()
    def handle_read_error(self):
        """Обработка ошибки чтения - установка статуса потери связи"""
        # Устанавливаем статус потери связи для обоих датчиков
        old_working1 = self.sensor_data[0]["working"]
        old_working2 = self.sensor_data[1]["working"]
        
        self.sensor_data[0]["working"] = False
        self.sensor_data[1]["working"] = False
        self.sensor_data[0]["temp"] = "ERROR"
        self.sensor_data[1]["temp"] = "ERROR"
        
        # Обновляем отображение
        self.update_display()
        
        # Сохраняем в Excel только если статус изменился
        if old_working1 or old_working2:
            self.save_to_excel_if_changed()
    
    @pyqtSlot(str)
    def process_line(self, line):
        """Обработка полученной строки"""
        # Если была ошибка чтения, сбрасываем флаг при успешном чтении
        if self.read_error_occurred:
            self.read_error_occurred = False
            self.reconnect_mode = False
            self.status_bar.showMessage("Связь восстановлена", 3000)
            
            # Восстанавливаем нормальный вид кнопки
            self.connect_btn.setText("🔌 Отключиться")
            self.connect_btn.setStyleSheet("""
                QPushButton {
                    font-size: 20px;
                    font-weight: bold;
                    padding: 10px 15px;
                    background-color: #e74c3c;
                    color: white;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
        
        # Парсим температуру
        if self.parse_temperature(line):
            # Сохраняем только если температура изменилась
            self.save_to_excel_if_changed()
        
        # Проверяем на отключение датчиков
        if any(word in line.lower() for word in ["not found", "no sensor", "failed", "отсутствует", "error"]):
            if self.check_sensor_error(line):
                # Сохраняем статус ошибки
                self.save_to_excel_if_changed()
        
        # Проверяем на изменение разрешения
        if "changed" in line.lower():
            self.parse_resolution(line)
    
    def parse_temperature(self, line):
        """Парсинг температуры, возвращает True если данные изменились"""
        # Ищем все числа с точкой в строке
        temperatures = re.findall(r'-?\d+\.\d+', line)
        
        if len(temperatures) >= 2:
            # Нашли две температуры
            old_temp1 = self.sensor_data[0]["temp"]
            old_temp2 = self.sensor_data[1]["temp"]
            
            self.sensor_data[0]["temp"] = temperatures[0]
            self.sensor_data[1]["temp"] = temperatures[1]
            self.sensor_data[0]["working"] = True
            self.sensor_data[1]["working"] = True
            
            # Проверяем изменилась ли температура
            changed = (old_temp1 != temperatures[0]) or (old_temp2 != temperatures[1])
            
            self.update_display()
            return changed
            
        elif len(temperatures) == 1:
            # Нашли одну температуру - предполагаем, что это датчик 0
            old_temp1 = self.sensor_data[0]["temp"]
            self.sensor_data[0]["temp"] = temperatures[0]
            self.sensor_data[0]["working"] = True
            changed = old_temp1 != temperatures[0]
            self.update_display()
            return changed
        
        return False
    
    def check_sensor_error(self, line):
        """Проверка ошибок датчиков, возвращает True если статус изменился"""
        old_working1 = self.sensor_data[0]["working"]
        old_working2 = self.sensor_data[1]["working"]
        
        if 's0' in line.lower():
            self.sensor_data[0]["working"] = False
            self.sensor_data[0]["temp"] = "ERROR"
            self.update_display()
            self.status_bar.showMessage("ДАТЧИК 1: НЕТ СВЯЗИ!", 5000)
            return old_working1 != False
        
        if 's1' in line.lower():
            self.sensor_data[1]["working"] = False
            self.sensor_data[1]["temp"] = "ERROR"
            self.update_display()
            self.status_bar.showMessage("ДАТЧИК 2: НЕТ СВЯЗИ!", 5000)
            return old_working2 != False
        
        return False
    
    def parse_resolution(self, line):
        """Парсинг изменения разрешения"""
        if 's0' in line.lower():
            if '9-bit' in line:
                self.sensor_data[0]["res"] = "9"
                self.sensor1_res_buttons["9"].setChecked(True)
            elif '10-bit' in line:
                self.sensor_data[0]["res"] = "10"
                self.sensor1_res_buttons["10"].setChecked(True)
            elif '11-bit' in line:
                self.sensor_data[0]["res"] = "11"
                self.sensor1_res_buttons["11"].setChecked(True)
            elif '12-bit' in line:
                self.sensor_data[0]["res"] = "12"
                self.sensor1_res_buttons["12"].setChecked(True)
        
        if 's1' in line.lower():
            if '9-bit' in line:
                self.sensor_data[1]["res"] = "9"
                self.sensor2_res_buttons["9"].setChecked(True)
            elif '10-bit' in line or '10 bit' in line:
                self.sensor_data[1]["res"] = "10"
                self.sensor2_res_buttons["10"].setChecked(True)
            elif '11-bit' in line or '11 bit' in line:
                self.sensor_data[1]["res"] = "11"
                self.sensor2_res_buttons["11"].setChecked(True)
            elif '12-bit' in line or '12 bit' in line:
                self.sensor_data[1]["res"] = "12"
                self.sensor2_res_buttons["12"].setChecked(True)
        
        self.update_display()
        
        # Сохраняем изменение разрешения
        self.save_to_excel_if_changed()
    
    def update_display(self):
        """Обновление отображения"""
        # Датчик 1
        temp1 = self.sensor_data[0]["temp"]
        working1 = self.sensor_data[0]["working"]
        res1 = self.sensor_data[0]["res"]
        
        self.sensor1_temp.setText(f"{temp1} °C")
        if working1:
            self.sensor1_status.setText(f"✓ Работает | {res1} бит")
            self.sensor1_status.setStyleSheet("font-size: 30px; color: #27ae60; font-weight: bold; padding: 10px;")
            self.sensor1_frame.setStyleSheet("border: 3px solid #27ae60; background-color: #f0f8ff;")
        else:
            self.sensor1_status.setText("✗ ПОТЕРЯ СИГНАЛА")
            self.sensor1_status.setStyleSheet("font-size: 30px; color: #e74c3c; font-weight: bold; padding: 10px;")
            self.sensor1_frame.setStyleSheet("border: 3px solid #e74c3c; background-color: #fff0f0;")
        
        # Датчик 2
        temp2 = self.sensor_data[1]["temp"]
        working2 = self.sensor_data[1]["working"]
        res2 = self.sensor_data[1]["res"]
        
        self.sensor2_temp.setText(f"{temp2} °C")
        if working2:
            self.sensor2_status.setText(f"✓ Работает | {res2} бит")
            self.sensor2_status.setStyleSheet("font-size: 30px; color: #27ae60; font-weight: bold; padding: 10px;")
            self.sensor2_frame.setStyleSheet("border: 3px solid #27ae60; background-color: #f0f8ff;")
        else:
            self.sensor2_status.setText("✗ ПОТЕРЯ СИГНАЛА")
            self.sensor2_status.setStyleSheet("font-size: 30px; color: #e74c3c; font-weight: bold; padding: 10px;")
            self.sensor2_frame.setStyleSheet("border: 3px solid #e74c3c; background-color: #fff0f0;")
    
    def send_command(self, cmd):
        """Отправка команды"""
        if self.is_connected and self.serial_port:
            try:
                self.serial_port.write(f"{cmd}\n".encode())
                self.status_bar.showMessage(f"Команда отправлена: '{cmd}'", 3000)
                
            except Exception as e:
                self.status_bar.showMessage(f"Ошибка отправки: {e}", 5000)
        else:
            self.status_bar.showMessage("Не подключено к порту!", 5000)
    
    def keyPressEvent(self, event):
        """Обработка нажатий клавиш"""
        if event.key() == Qt.Key_F11:
            # Переключение полноэкранного режима
            if self.isFullScreen():
                self.showNormal()
            else:
                self.showFullScreen()
        elif event.key() == Qt.Key_Escape:
            # Выход из полноэкранного режима
            if self.isFullScreen():
                self.showNormal()
        else:
            super().keyPressEvent(event)
    
    def closeEvent(self, event):
        """Обработка закрытия окна"""
        self.disconnect()
        event.accept()

def main():
    # Убираем консольное окно на Windows
    if sys.platform == "win32":
        import ctypes
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    
    app = QApplication(sys.argv)
    
    # Устанавливаем стиль
    app.setStyle("Fusion")
    
    # Создаем и показываем окно
    window = DS18B20Monitor()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()